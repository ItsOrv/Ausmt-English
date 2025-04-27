import os
import pandas as pd
import logging
from dotenv import load_dotenv
from utils.config import Config
import openpyxl
from openpyxl.utils import column_index_from_string

# Load environment variables
load_dotenv()

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class ExcelHandler:
    def __init__(self):
        """Initialize Excel handler with path from environment variables."""
        self.excel_path = Config.EXCEL_PATH
        # Store column indices (convert from Excel column letters to numeric indices)
        self.column_indices = {
            'email': column_index_from_string(Config.EMAIL_COLUMN) - 1,
            'first_name': column_index_from_string(Config.FIRST_NAME_COLUMN) - 1,
            'last_name': column_index_from_string(Config.LAST_NAME_COLUMN) - 1,
            'father_name': column_index_from_string(Config.FATHER_NAME_COLUMN) - 1,
            'faculty': column_index_from_string(Config.FACULTY_COLUMN) - 1,
            'major': column_index_from_string(Config.MAJOR_COLUMN) - 1,
            'education_level': column_index_from_string(Config.EDUCATION_LEVEL_COLUMN) - 1,
            'phone': column_index_from_string(Config.PHONE_COLUMN) - 1
        }
        self._validate_excel_file()
    
    def _validate_excel_file(self):
        """Validate that the Excel file exists, if not create a sample one."""
        if not os.path.exists(self.excel_path):
            # Create directory if it doesn't exist
            os.makedirs(os.path.dirname(self.excel_path), exist_ok=True)
            
            # Create a workbook with sample data
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Create sample data with cells in the correct columns
            sample_data = [
                # Student 1
                ('9512345', 'محمد', 'احمدی', 'علی', 'مهندسی', 'کامپیوتر', 'کارشناسی', 'mohammad@example.com', '09123456789'),
                # Student 2
                ('9623456', 'علی', 'محمدی', 'حسن', 'علوم پایه', 'فیزیک', 'کارشناسی ارشد', 'ali@example.com', '09123456788'),
                # Student 3
                ('9734567', 'فاطمه', 'حسینی', 'محمود', 'علوم انسانی', 'روانشناسی', 'کارشناسی', 'fateme@example.com', '09123456787'),
                # Student 4
                ('9845678', 'زهرا', 'رضایی', 'رضا', 'مهندسی', 'برق', 'دکتری', 'zahra@example.com', '09123456786'),
                # Student 5
                ('9956789', 'امیر', 'کریمی', 'مجید', 'پزشکی', 'پزشکی عمومی', 'دکتری', 'amir@example.com', '09123456785')
            ]
            
            # Add student_id in the first column (assuming it's column 0)
            student_id_col = 0
            for row_idx, student in enumerate(sample_data, start=1):
                ws.cell(row=row_idx, column=student_id_col+1, value=student[0])
                
                # Add other data in their respective columns
                ws.cell(row=row_idx, column=self.column_indices['email']+1, value=student[7])
                ws.cell(row=row_idx, column=self.column_indices['first_name']+1, value=student[1])
                ws.cell(row=row_idx, column=self.column_indices['last_name']+1, value=student[2])
                ws.cell(row=row_idx, column=self.column_indices['father_name']+1, value=student[3])
                ws.cell(row=row_idx, column=self.column_indices['faculty']+1, value=student[4])
                ws.cell(row=row_idx, column=self.column_indices['major']+1, value=student[5])
                ws.cell(row=row_idx, column=self.column_indices['education_level']+1, value=student[6])
                ws.cell(row=row_idx, column=self.column_indices['phone']+1, value=student[8])
            
            # Save the workbook
            wb.save(self.excel_path)
            logger.info(f"Created sample Excel file at {self.excel_path}")
    
    def check_student_id(self, student_id):
        """
        Check if a student ID exists in the Excel file.
        
        Args:
            student_id (str): Student ID to check
            
        Returns:
            tuple: (exists, first_name, last_name) or (False, None, None) if not found
        """
        try:
            # Load the Excel file
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            
            # Convert student_id to string to ensure proper comparison
            student_id = str(student_id)
            
            # Find the student
            found = False
            first_name = None
            last_name = None
            
            # Debug: Print all student IDs in the Excel file
            logger.info(f"Looking for student ID: {student_id}")
            logger.info("Student IDs in the Excel file:")
            for row in ws.iter_rows(min_row=1, max_col=1, max_row=ws.max_row):
                cell_value = str(row[0].value) if row[0].value is not None else ""
                logger.info(f"  {cell_value}")
                
            # Assuming student_id is in column A (index 0)
            for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
                cell_value = str(row[0].value) if row[0].value is not None else ""
                logger.info(f"Comparing '{cell_value}' with '{student_id}'")
                
                if cell_value == student_id:
                    logger.info(f"Match found! Fetching first_name from column index {self.column_indices['first_name']}")
                    logger.info(f"Fetching last_name from column index {self.column_indices['last_name']}")
                    
                    first_name = row[self.column_indices['first_name']].value
                    last_name = row[self.column_indices['last_name']].value
                    logger.info(f"Found student: {first_name} {last_name}")
                    found = True
                    break
            
            return found, first_name, last_name
                
        except Exception as e:
            logger.error(f"Error checking student ID: {e}")
            return False, None, None 